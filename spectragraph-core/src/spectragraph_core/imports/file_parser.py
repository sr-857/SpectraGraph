"""
File parsing utilities for entity imports.
Handles CSV, TXT, and XLSX file formats.
Each row represents ONE entity with all columns as data properties.
"""

import csv
import io
from dataclasses import dataclass
from typing import List, Dict, Any, BinaryIO, Union, Optional
from pathlib import Path

try:
    from openpyxl import load_workbook

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .type_matcher import detect_entity_type_from_row, get_primary_identifier


@dataclass
class EntityPreview:
    """Preview of a single entity to be imported."""

    row_index: int
    """Index of the row (0-based)."""

    data: Dict[str, Any]
    """All column data as key-value pairs."""

    detected_type: str
    """Detected entity type (e.g., "Email", "Domain")."""

    primary_value: str
    """Primary identifier value for the entity."""

    confidence: str
    """Confidence level: high, medium, low."""


@dataclass
class FileParseResult:
    """Result of parsing an import file."""

    entities: List[EntityPreview]
    """List of detected entities from the file."""

    total_entities: int
    """Total number of entities (rows)."""

    type_distribution: Dict[str, int]
    """Distribution of detected types: {type_name: count}."""

    columns: List[str]
    """Column names found in the file (for reference)."""


def parse_file(
    file_content: Union[bytes, BinaryIO],
    filename: str,
    max_preview_rows: int = 100,
) -> FileParseResult:
    """
    Parse an uploaded file and analyze its contents.
    Each row is treated as a single entity.

    Args:
        file_content: File content as bytes or file-like object
        filename: Original filename (used to determine file type)
        max_preview_rows: Maximum number of rows to parse for preview

    Returns:
        FileParseResult containing parsed entities and analysis

    Raises:
        ValueError: If file format is unsupported or parsing fails
    """
    file_ext = Path(filename).suffix.lower()

    # Convert file content to bytes if it's a file-like object
    if hasattr(file_content, "read"):
        file_bytes = file_content.read()
    else:
        file_bytes = file_content

    # Parse based on file extension
    if file_ext == ".csv":
        return _parse_csv(file_bytes, max_preview_rows)
    elif file_ext == ".txt":
        return _parse_txt(file_bytes, max_preview_rows)
    elif file_ext in [".xlsx", ".xls"]:
        if not OPENPYXL_AVAILABLE:
            raise ValueError(
                "openpyxl library is required for Excel file support. "
                "Install with: pip install openpyxl"
            )
        return _parse_excel(file_bytes, max_preview_rows)
    else:
        raise ValueError(f"Unsupported file format: {file_ext}")


def _parse_csv(
    file_bytes: bytes,
    max_preview_rows: int,
) -> FileParseResult:
    """Parse a CSV file where each row is an entity."""
    try:
        # Try to decode as UTF-8, fall back to latin-1
        try:
            text_content = file_bytes.decode("utf-8")
        except UnicodeDecodeError:
            text_content = file_bytes.decode("latin-1")

        # Parse CSV
        csv_reader = csv.DictReader(io.StringIO(text_content))

        # Extract columns
        columns = csv_reader.fieldnames or []
        if not columns:
            raise ValueError("No columns found in CSV file")

        # Parse rows as entities
        entities: List[EntityPreview] = []
        for row_idx, row in enumerate(csv_reader):
            if row_idx >= max_preview_rows:
                break

            # Skip empty rows
            if not any(v and str(v).strip() for v in row.values()):
                continue

            entity = _create_entity_preview(row_idx, row, columns)
            if entity:
                entities.append(entity)

        return _create_result(entities, columns)

    except Exception as e:
        raise ValueError(f"Failed to parse CSV file: {str(e)}")


def _parse_txt(
    file_bytes: bytes,
    max_preview_rows: int,
) -> FileParseResult:
    """Parse a TXT file where each line is an entity with a single 'value' column."""
    try:
        # Try to decode as UTF-8, fall back to latin-1
        try:
            text_content = file_bytes.decode("utf-8")
        except UnicodeDecodeError:
            text_content = file_bytes.decode("latin-1")

        # Split by lines and filter empty lines
        lines = [line.strip() for line in text_content.split("\n")]
        lines = [line for line in lines if line]

        if not lines:
            raise ValueError("File is empty")

        # For TXT files, single column named "value"
        columns = ["value"]

        # Convert lines to entities
        entities: List[EntityPreview] = []
        for row_idx, line in enumerate(lines[:max_preview_rows]):
            row_data = {"value": line}
            entity = _create_entity_preview(row_idx, row_data, columns)
            if entity:
                entities.append(entity)

        return _create_result(entities, columns)

    except Exception as e:
        raise ValueError(f"Failed to parse TXT file: {str(e)}")


def _parse_excel(
    file_bytes: bytes,
    max_preview_rows: int,
) -> FileParseResult:
    """Parse an Excel file where each row is an entity."""
    try:
        # Load workbook from bytes
        workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)

        # Use the first sheet
        sheet = workbook.active
        if sheet is None:
            raise ValueError("No active sheet found in Excel file")

        # Read all rows
        all_rows = list(sheet.iter_rows(values_only=True))

        if not all_rows:
            raise ValueError("Excel file is empty")

        # First row is header
        header_row = all_rows[0]
        data_rows = all_rows[1:]

        # Convert header to strings and handle None values
        columns = [
            str(cell) if cell is not None else f"Column_{i+1}"
            for i, cell in enumerate(header_row)
        ]

        # Convert data rows to entities
        entities: List[EntityPreview] = []
        for row_idx, row_values in enumerate(data_rows[:max_preview_rows]):
            row_dict = {}
            for i, value in enumerate(row_values):
                if i < len(columns):
                    row_dict[columns[i]] = str(value) if value is not None else ""

            # Skip empty rows
            if not any(v and str(v).strip() for v in row_dict.values()):
                continue

            entity = _create_entity_preview(row_idx, row_dict, columns)
            if entity:
                entities.append(entity)

        workbook.close()

        return _create_result(entities, columns)

    except Exception as e:
        raise ValueError(f"Failed to parse Excel file: {str(e)}")


def _create_entity_preview(
    row_idx: int, row_data: Dict[str, Any], columns: List[str]
) -> Optional[EntityPreview]:
    """
    Create an EntityPreview from a row of data.

    Args:
        row_idx: Row index
        row_data: Dictionary of column: value pairs
        columns: List of column names

    Returns:
        EntityPreview or None if row is invalid
    """
    # Filter out empty values
    data = {k: v for k, v in row_data.items() if v and str(v).strip()}

    if not data:
        return None

    # Detect entity type
    detected_type = detect_entity_type_from_row(data)
    if not detected_type:
        detected_type = "Unknown"

    # Get primary identifier
    primary_value = get_primary_identifier(data, detected_type)
    if not primary_value:
        # Fallback to first value
        primary_value = str(next(iter(data.values())))

    # Determine confidence
    confidence = _determine_confidence(detected_type, data)

    return EntityPreview(
        row_index=row_idx,
        data=data,
        detected_type=detected_type,
        primary_value=primary_value,
        confidence=confidence,
    )


def _determine_confidence(entity_type: str, data: Dict[str, Any]) -> str:
    """
    Determine confidence level for entity type detection.

    Args:
        entity_type: Detected entity type
        data: Row data

    Returns:
        "high", "medium", or "low"
    """
    if entity_type == "Unknown":
        return "low"

    # Normalize keys
    keys = [k.lower().strip() for k in data.keys()]

    # Check if column names strongly indicate the type
    type_keywords = {
        "Email": ["email", "mail", "e-mail"],
        "Ip": ["ip", "ip_address", "ipv4", "ipv6"],
        "Domain": ["domain", "hostname"],
        "Website": ["url", "website", "web"],
        "Phone": ["phone", "telephone", "tel", "mobile"],
        "ASN": ["asn", "as_number"],
        "Username": ["username", "handle", "user"],
        "Organization": ["organization", "org", "company"],
        "Individual": ["name", "person", "individual"],
    }

    keywords = type_keywords.get(entity_type, [])
    if any(keyword in key for key in keys for keyword in keywords):
        return "high"

    # Medium confidence if type was detected but no explicit column names
    if entity_type != "Unknown":
        return "medium"

    return "low"


def _create_result(
    entities: List[EntityPreview], columns: List[str]
) -> FileParseResult:
    """Create FileParseResult from entities."""
    # Calculate type distribution
    type_distribution: Dict[str, int] = {}
    for entity in entities:
        type_name = entity.detected_type
        type_distribution[type_name] = type_distribution.get(type_name, 0) + 1

    return FileParseResult(
        entities=entities,
        total_entities=len(entities),
        type_distribution=type_distribution,
        columns=columns,
    )
